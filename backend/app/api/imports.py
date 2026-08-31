"""Master data import endpoints — bulk upload employees and payroll history from Excel/CSV.

Used for initial onboarding: tax accountant exports data from WeHaGo T (or any spreadsheet)
and uploads it here to populate Employee master + PayrollEntry history.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user, get_db
from app.models import (
    Client,
    CollectionEvent,
    CollectionSession,
    CollectionSessionStatus,
    Employee,
    EmploymentStatus,
    IncomeType,
    MatchStatus,
    MonthlyFiling,
    MonthlyFilingStatus,
    PayrollEntry,
    User,
)
from app.schemas.clients import EmployeeOut
from app.schemas.imports import ImportEmployeeResult, ImportPayrollResult
from app.services.crypto import encrypt_rrn, mask_rrn
from app.services.tax_calc import calculate_withholding_tax, income_type_to_a_code

router = APIRouter()

# ---------------------------------------------------------------------------
# Column alias mapping — covers WeHaGo T export + common free-form headers
# ---------------------------------------------------------------------------

_EMPLOYEE_ALIASES: dict[str, list[str]] = {
    "name": ["성명", "이름", "직원명", "name", "사원명"],
    "employee_code": ["사번", "사원번호", "직원코드", "employee_code", "코드"],
    "rrn": ["주민등록번호", "주민번호", "rrn", "주민", "생년월일"],
    "hired_at": ["입사일", "입사년월일", "입사일자", "hired_at"],
    "resigned_at": ["퇴사일", "퇴사년월일", "퇴사일자", "resigned_at"],
    "income_type": ["소득구분", "소득종류", "소득구분코드", "income_type"],
}

_PAYROLL_ALIASES: dict[str, list[str]] = {
    "name": ["성명", "이름", "직원명", "name", "사원명"],
    "employee_code": ["사번", "사원번호", "직원코드", "employee_code", "코드"],
    "total_amount": ["총지급액", "지급액", "급여", "급여액", "total_amount", "지급총액"],
    "non_taxable": ["비과세", "비과세소득", "비과세액", "non_taxable"],
    "income_type": ["소득구분", "소득종류", "소득구분코드", "income_type"],
}


def _map_columns(headers: list[str], aliases: dict[str, list[str]]) -> dict[str, int]:
    """Map logical field names to column indices using alias lookup."""
    mapping: dict[str, int] = {}
    normalized = [h.strip().lower().replace(" ", "") for h in headers]
    for field, candidates in aliases.items():
        for alias in candidates:
            alias_norm = alias.strip().lower().replace(" ", "")
            for idx, hdr in enumerate(normalized):
                if hdr == alias_norm and field not in mapping:
                    mapping[field] = idx
                    break
            if field in mapping:
                break
    return mapping


def _read_rows(content: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    """Read an Excel or CSV file into header + data rows."""
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_raw = next(rows_iter, None)
        if not header_raw:
            return [], []
        headers = [str(c).strip() if c is not None else "" for c in header_raw]
        data = []
        for row in rows_iter:
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                data.append(cells)
        return headers, data
    elif lower.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        header_raw = next(reader, None)
        if not header_raw:
            return [], []
        headers = [h.strip() for h in header_raw]
        data = [[c.strip() for c in row] for row in reader if any(c.strip() for c in row)]
        return headers, data
    else:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"지원하지 않는 파일 형식: {filename} (.xlsx, .csv만 지원)",
        )


def _parse_date(value: str) -> date | None:
    """Try common date formats."""
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            from datetime import datetime
            return datetime.strptime(value.replace(" ", ""), fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(value: str) -> int:
    """Parse Korean currency strings: '3,000,000' or '300만원' → int."""
    if not value:
        return 0
    cleaned = re.sub(r"[^\d.만]", "", value)
    if "만" in cleaned:
        num_part = cleaned.replace("만", "")
        try:
            return int(float(num_part) * 10_000)
        except ValueError:
            return 0
    try:
        return int(float(cleaned))
    except ValueError:
        return 0


def _parse_income_type(value: str) -> IncomeType:
    """Map Korean income type labels to enum."""
    v = value.strip().upper()
    mapping = {
        "근로": IncomeType.WAGE, "WAGE": IncomeType.WAGE,
        "사업": IncomeType.BUSINESS, "BUSINESS": IncomeType.BUSINESS,
        "기타": IncomeType.OTHER, "OTHER": IncomeType.OTHER,
        "일용": IncomeType.DAILY, "DAILY": IncomeType.DAILY,
        "퇴직": IncomeType.RETIREMENT, "RETIREMENT": IncomeType.RETIREMENT,
    }
    return mapping.get(v, IncomeType.WAGE)


def _get_cell(row: list[str], mapping: dict[str, int], field: str) -> str:
    idx = mapping.get(field)
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.post("/{client_id}/import-employees", response_model=ImportEmployeeResult)
async def import_employees(
    client_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> ImportEmployeeResult:
    """직원 마스터 일괄 임포트 (위하고T 인적사항 엑셀 or 자유 양식)."""
    client = await db.get(Client, client_id)
    if not client or client.tax_office_id != user.tax_office_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Client not found")

    content = await file.read()
    if not content:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "빈 파일입니다")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "10MB 초과")

    headers, data = _read_rows(content, file.filename or "upload.xlsx")
    col_map = _map_columns(headers, _EMPLOYEE_ALIASES)

    if "name" not in col_map:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"'성명' 또는 '이름' 컬럼을 찾을 수 없습니다. 헤더: {headers}",
        )

    # Load existing employees for dedup
    existing = (
        await db.execute(select(Employee).where(Employee.client_id == client_id))
    ).scalars().all()
    by_name: dict[str, Employee] = {e.name: e for e in existing}

    created, updated, skipped = 0, 0, 0
    errors: list[str] = []
    result_employees: list[Employee] = []

    for i, row in enumerate(data, start=2):  # row 2 = first data row in Excel
        name = _get_cell(row, col_map, "name")
        if not name:
            skipped += 1
            continue

        code = _get_cell(row, col_map, "employee_code") or None
        rrn_raw = _get_cell(row, col_map, "rrn") or None
        hired = _parse_date(_get_cell(row, col_map, "hired_at"))
        resigned = _parse_date(_get_cell(row, col_map, "resigned_at"))
        income_type_str = _get_cell(row, col_map, "income_type")

        try:
            rrn_encrypted = encrypt_rrn(rrn_raw) if rrn_raw else None
            rrn_last4 = rrn_raw.replace("-", "")[-4:] if rrn_raw and len(rrn_raw.replace("-", "")) >= 4 else None
        except Exception:  # noqa: BLE001
            errors.append(f"행 {i}: {name} — 주민번호 암호화 실패")
            rrn_encrypted = None
            rrn_last4 = None

        emp_status = EmploymentStatus.RESIGNED if resigned else (
            EmploymentStatus.ACTIVE if rrn_raw else EmploymentStatus.PENDING
        )

        if name in by_name:
            # Update existing
            emp = by_name[name]
            if code:
                emp.employee_code = code
            if rrn_encrypted:
                emp.rrn_encrypted = rrn_encrypted
                emp.rrn_last4 = rrn_last4
            if hired:
                emp.hired_at = hired
            if resigned:
                emp.resigned_at = resigned
                emp.status = EmploymentStatus.RESIGNED
            elif emp.status == EmploymentStatus.PENDING and rrn_raw:
                emp.status = EmploymentStatus.ACTIVE
            updated += 1
            result_employees.append(emp)
        else:
            emp = Employee(
                client_id=client_id,
                name=name,
                employee_code=code,
                rrn_encrypted=rrn_encrypted,
                rrn_last4=rrn_last4,
                hired_at=hired,
                resigned_at=resigned,
                status=emp_status,
            )
            db.add(emp)
            by_name[name] = emp
            created += 1
            result_employees.append(emp)

    await db.commit()
    for emp in result_employees:
        await db.refresh(emp)

    return ImportEmployeeResult(
        total_rows=len(data),
        created=created,
        updated=updated,
        skipped=skipped,
        errors=errors,
        employees=[EmployeeOut.model_validate(e) for e in result_employees],
    )


@router.post("/{client_id}/import-payroll", response_model=ImportPayrollResult)
async def import_payroll(
    client_id: str,
    period: str = Query(..., pattern=r"^\d{4}-\d{2}$", description="지급년월 (YYYY-MM)"),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> ImportPayrollResult:
    """전월 급여 이력 일괄 임포트."""
    client = await db.get(Client, client_id)
    if not client or client.tax_office_id != user.tax_office_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Client not found")

    content = await file.read()
    if not content:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "빈 파일입니다")

    headers, data = _read_rows(content, file.filename or "upload.xlsx")
    col_map = _map_columns(headers, _PAYROLL_ALIASES)

    if "name" not in col_map and "employee_code" not in col_map:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"'성명' 또는 '사번' 컬럼을 찾을 수 없습니다. 헤더: {headers}",
        )
    if "total_amount" not in col_map:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"'총지급액' 또는 '급여' 컬럼을 찾을 수 없습니다. 헤더: {headers}",
        )

    # Load employees for matching
    employees = (
        await db.execute(select(Employee).where(Employee.client_id == client_id))
    ).scalars().all()
    by_name = {e.name: e for e in employees}
    by_code = {e.employee_code: e for e in employees if e.employee_code}

    # Get or create MonthlyFiling
    filing = (
        await db.execute(
            select(MonthlyFiling).where(
                MonthlyFiling.tax_office_id == user.tax_office_id,
                MonthlyFiling.period == period,
            )
        )
    ).scalar_one_or_none()
    if not filing:
        filing = MonthlyFiling(
            tax_office_id=user.tax_office_id,
            period=period,
            status=MonthlyFilingStatus.APPROVED,
        )
        db.add(filing)
        await db.flush()

    # Create a CollectionSession for the import
    from app.services.crypto import random_token
    session = CollectionSession(
        monthly_filing_id=filing.id,
        client_id=client_id,
        status=CollectionSessionStatus.APPROVED,
        request_token=random_token(32),
    )
    db.add(session)
    await db.flush()

    db.add(CollectionEvent(
        session_id=session.id,
        event_type="IMPORT_PAYROLL",
        channel="import",
        raw_text=f"엑셀 임포트: {file.filename}, {len(data)} 행",
    ))

    matched_count, unmatched_count, created_count = 0, 0, 0
    errors: list[str] = []

    for i, row in enumerate(data, start=2):
        name = _get_cell(row, col_map, "name")
        code = _get_cell(row, col_map, "employee_code")
        total_raw = _get_cell(row, col_map, "total_amount")
        non_tax_raw = _get_cell(row, col_map, "non_taxable")
        income_type_str = _get_cell(row, col_map, "income_type")

        total_amount = _parse_amount(total_raw)
        if total_amount <= 0:
            errors.append(f"행 {i}: 총지급액이 0 이하 — 건너뜀")
            continue

        non_taxable = _parse_amount(non_tax_raw)
        income_type = _parse_income_type(income_type_str)

        # Match to employee
        emp = None
        if code and code in by_code:
            emp = by_code[code]
        elif name and name in by_name:
            emp = by_name[name]

        if emp:
            matched_count += 1
            match_status = MatchStatus.MATCHED
        else:
            unmatched_count += 1
            match_status = MatchStatus.AMBIGUOUS
            errors.append(f"행 {i}: '{name or code}' — 직원 매칭 실패")

        # 비과세는 상용근로(WAGE)에만 존재하고, 총지급액에 이미 포함된 값이다.
        if income_type != IncomeType.WAGE:
            non_taxable = 0
        else:
            non_taxable = min(non_taxable, total_amount)
        taxable = total_amount - non_taxable
        biz_code = emp.business_type_code if emp and income_type == IncomeType.BUSINESS else None
        tax = calculate_withholding_tax(
            income_type, taxable, dependents=1, business_type_code=biz_code,
        )
        a_code = income_type_to_a_code(income_type, is_corporation=client.is_corporation)
        salary_amt = total_amount if income_type == IncomeType.WAGE else None
        bonus_amt = 0 if income_type == IncomeType.WAGE else None

        entry = PayrollEntry(
            monthly_filing_id=filing.id,
            collection_session_id=session.id,
            client_id=client_id,
            employee_id=emp.id if emp else None,
            raw_name=name or code or f"행{i}",
            income_type=income_type,
            a_code=a_code,
            business_type_code=biz_code,
            total_amount=total_amount,
            salary_amount=salary_amt,
            bonus_amount=bonus_amt,
            non_taxable=non_taxable,
            taxable=taxable,
            income_tax=tax.income_tax,
            local_tax=tax.local_tax,
            match_status=match_status,
            approved=bool(emp),
        )
        db.add(entry)
        created_count += 1

    filing.total_clients = max(filing.total_clients, 1)
    filing.total_entries = (filing.total_entries or 0) + created_count
    await db.commit()

    return ImportPayrollResult(
        period=period,
        total_rows=len(data),
        matched=matched_count,
        unmatched=unmatched_count,
        created_entries=created_count,
        errors=errors,
    )
