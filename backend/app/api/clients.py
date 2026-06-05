"""Client (거래처) + Employee endpoints."""

import io
import logging

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user, get_db
from app.models import (
    Client,
    ClientPayrollDefault,
    Employee,
    EmploymentStatus,
    MonthlyFiling,
    MonthlyFilingStatus,
    TaxOffice,
    User,
)
from app.schemas.clients import (
    ChannelAttempt,
    ClientCreate,
    ClientInviteResult,
    ClientOut,
    ClientUpdate,
    EmployeeCreate,
    EmployeeOut,
    PayrollDefaultOut,
    PayrollDefaultUpdate,
)
from app.services.crypto import encrypt_rrn, mask_rrn
from app.services.invite import get_or_create_session, send_invite_to_client
from app.services.tax_calc import (
    DEFAULT_EI_RATE,
    DEFAULT_HI_RATE,
    DEFAULT_LTC_RATE_OF_HI,
    DEFAULT_NPS_RATE,
)

logger = logging.getLogger(__name__)


# 신규 거래처가 자동 추가될 수 있는 활성 filing 상태 (제출·완료된 filing은 제외)
_ACTIVE_FILING_STATUSES = (
    MonthlyFilingStatus.DRAFT,
    MonthlyFilingStatus.COLLECTING,
    MonthlyFilingStatus.REVIEWING,
)

router = APIRouter()


@router.get("", response_model=list[ClientOut])
async def list_clients(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[Client]:
    rows = (
        await db.execute(
            select(Client).where(Client.tax_office_id == user.tax_office_id).order_by(Client.business_name)
        )
    ).scalars().all()
    return list(rows)


@router.post("", response_model=ClientOut, status_code=status.HTTP_201_CREATED)
async def create_client(
    payload: ClientCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Client:
    client = Client(
        tax_office_id=user.tax_office_id,
        business_name=payload.business_name,
        business_number=payload.business_number,
        representative=payload.representative,
        contact_phone=payload.contact_phone,
        contact_email=str(payload.contact_email) if payload.contact_email else None,
        is_corporation=payload.is_corporation,
    )
    db.add(client)
    await db.flush()  # client.id 확보

    # 활성 월별신고가 있으면 자동으로 수집 세션 생성 → 신고 대시보드에 즉시 노출
    active_filing = (
        await db.execute(
            select(MonthlyFiling)
            .where(
                MonthlyFiling.tax_office_id == user.tax_office_id,
                MonthlyFiling.status.in_(_ACTIVE_FILING_STATUSES),
            )
            .order_by(MonthlyFiling.period.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
    if active_filing is not None:
        await get_or_create_session(db, active_filing, client, ttl_days=30)

    await db.commit()
    await db.refresh(client)
    return client


@router.post("/bulk-upload", response_model=list[ClientOut], status_code=status.HTTP_201_CREATED)
async def bulk_upload_clients(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[Client]:
    """엑셀/CSV 파일로 거래처 일괄 등록.

    필수 컬럼: 상호 (또는 business_name, 사업자명)
    선택 컬럼: 사업자번호, 대표자, 전화번호, 이메일, 법인여부
    """
    content = await file.read()
    if not content:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "빈 파일입니다")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "10MB 초과")

    fname = (file.filename or "").lower()
    try:
        if fname.endswith(".csv"):
            import csv as csv_mod
            text = content.decode("utf-8-sig")
            reader = csv_mod.DictReader(io.StringIO(text))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(status.HTTP_400_BAD_REQUEST, "시트를 찾을 수 없습니다")
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {h: (str(v).strip() if v is not None else "") for h, v in zip(headers, row)}
                rows.append(row_dict)
            wb.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"파일 파싱 실패: {e}") from e

    if not rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "데이터 행이 없습니다")

    # Column name mapping (flexible)
    _COL_MAP = {
        "상호": "business_name", "사업자명": "business_name", "business_name": "business_name",
        "거래처명": "business_name", "회사명": "business_name",
        "사업자번호": "business_number", "business_number": "business_number",
        "대표자": "representative", "대표자명": "representative", "representative": "representative",
        "전화번호": "contact_phone", "연락처": "contact_phone", "휴대폰": "contact_phone",
        "contact_phone": "contact_phone", "phone": "contact_phone",
        "이메일": "contact_email", "email": "contact_email", "contact_email": "contact_email",
        "법인": "is_corporation", "법인여부": "is_corporation", "is_corporation": "is_corporation",
    }

    def _map_row(raw: dict) -> dict:
        mapped: dict = {}
        for k, v in raw.items():
            norm = _COL_MAP.get(k.strip())
            if norm and v:
                mapped[norm] = v
        return mapped

    # Get active filing for auto-session creation
    active_filing = (
        await db.execute(
            select(MonthlyFiling)
            .where(
                MonthlyFiling.tax_office_id == user.tax_office_id,
                MonthlyFiling.status.in_(_ACTIVE_FILING_STATUSES),
            )
            .order_by(MonthlyFiling.period.desc())
            .limit(1)
        )
    ).scalar_one_or_none()

    created: list[Client] = []
    for raw_row in rows:
        m = _map_row(raw_row)
        biz_name = m.get("business_name", "").strip()
        if not biz_name:
            continue

        is_corp = False
        is_corp_raw = m.get("is_corporation", "").strip().lower()
        if is_corp_raw in ("true", "1", "yes", "y", "법인", "○"):
            is_corp = True

        client = Client(
            tax_office_id=user.tax_office_id,
            business_name=biz_name,
            business_number=m.get("business_number") or None,
            representative=m.get("representative") or None,
            contact_phone=m.get("contact_phone") or None,
            contact_email=m.get("contact_email") or None,
            is_corporation=is_corp,
        )
        db.add(client)
        await db.flush()

        if active_filing is not None:
            await get_or_create_session(db, active_filing, client, ttl_days=30)

        created.append(client)

    if not created:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "유효한 거래처 데이터가 없습니다 (상호 컬럼 필수)")

    await db.commit()
    for c in created:
        await db.refresh(c)

    logger.info("Bulk uploaded %d clients for tax_office %s", len(created), user.tax_office_id)
    return created


@router.get("/{client_id}", response_model=ClientOut)
async def get_client(
    client_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Client:
    client = await db.get(Client, client_id)
    if not client or client.tax_office_id != user.tax_office_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Client not found")
    return client


@router.patch("/{client_id}", response_model=ClientOut)
async def update_client(
    client_id: str,
    payload: ClientUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Client:
    client = await db.get(Client, client_id)
    if not client or client.tax_office_id != user.tax_office_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Client not found")
    patch = payload.model_dump(exclude_unset=True)
    for field_name, value in patch.items():
        if field_name == "contact_email" and value is not None:
            value = str(value)
        setattr(client, field_name, value)
    await db.commit()
    await db.refresh(client)
    return client


@router.post("/{client_id}/invite", response_model=ClientInviteResult)
async def invite_client(
    client_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> ClientInviteResult:
    """단일 거래처에 초대장 발송 — 최신 신고 기간 기준 알림톡/SMS/이메일 발송."""
    client = await db.get(Client, client_id)
    if not client or client.tax_office_id != user.tax_office_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Client not found")

    if not client.contact_phone and not client.contact_email:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "연락처(전화번호 또는 이메일)를 먼저 입력해주세요.",
        )

    filing = (
        await db.execute(
            select(MonthlyFiling)
            .where(MonthlyFiling.tax_office_id == user.tax_office_id)
            .order_by(MonthlyFiling.period.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
    if filing is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "원천세 신고서를 먼저 생성해주세요.",
        )

    office = await db.get(TaxOffice, user.tax_office_id)
    office_name = office.name if office else "세무사사무소"

    _, accepted, attempts = await send_invite_to_client(db, filing, client, office_name)
    await db.commit()

    return ClientInviteResult(
        sent=bool(accepted),
        channels=accepted,
        attempts=[
            ChannelAttempt(channel=a.channel, accepted=a.accepted, error=a.error)
            for a in attempts
        ],
        filing_period=filing.period,
        detail=None if accepted else "모든 채널 발송 실패 — 연락처 정보를 확인해주세요.",
    )


@router.get("/{client_id}/employees", response_model=list[EmployeeOut])
async def list_employees(
    client_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[Employee]:
    client = await db.get(Client, client_id)
    if not client or client.tax_office_id != user.tax_office_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Client not found")
    rows = (
        await db.execute(select(Employee).where(Employee.client_id == client_id).order_by(Employee.name))
    ).scalars().all()
    return list(rows)


# ─── 거래처별 지급항목·4대보험 기본 세팅 (plan.md 3.8) ───


_PERCENT_TO_RATE = 100.0  # UI는 백분율(4.5), DB는 실수(0.045)


def _system_defaults_response() -> PayrollDefaultOut:
    """레코드가 없을 때 시스템 기본값으로 채운 응답."""
    return PayrollDefaultOut(
        nps_rate_percent=DEFAULT_NPS_RATE * _PERCENT_TO_RATE,
        hi_rate_percent=DEFAULT_HI_RATE * _PERCENT_TO_RATE,
        ltc_rate_percent=DEFAULT_LTC_RATE_OF_HI * _PERCENT_TO_RATE,
        ei_rate_percent=DEFAULT_EI_RATE * _PERCENT_TO_RATE,
        system_nps_rate_percent=DEFAULT_NPS_RATE * _PERCENT_TO_RATE,
        system_hi_rate_percent=DEFAULT_HI_RATE * _PERCENT_TO_RATE,
        system_ltc_rate_percent=DEFAULT_LTC_RATE_OF_HI * _PERCENT_TO_RATE,
        system_ei_rate_percent=DEFAULT_EI_RATE * _PERCENT_TO_RATE,
    )


def _to_response(row: ClientPayrollDefault) -> PayrollDefaultOut:
    """DB row → API. null 요율은 시스템 기본 요율로 채워서 반환."""

    def pct(val, default_rate: float) -> float:
        return float(val) * _PERCENT_TO_RATE if val is not None else default_rate * _PERCENT_TO_RATE

    return PayrollDefaultOut(
        meal_default=row.meal_default,
        car_default=row.car_default,
        childcare_default=row.childcare_default,
        apply_national_pension=row.apply_national_pension,
        apply_health_insurance=row.apply_health_insurance,
        apply_employment_insurance=row.apply_employment_insurance,
        apply_longterm_care=row.apply_longterm_care,
        nps_rate_percent=pct(row.nps_rate_override, DEFAULT_NPS_RATE),
        hi_rate_percent=pct(row.hi_rate_override, DEFAULT_HI_RATE),
        ltc_rate_percent=pct(row.ltc_rate_override, DEFAULT_LTC_RATE_OF_HI),
        ei_rate_percent=pct(row.ei_rate_override, DEFAULT_EI_RATE),
        note=row.note,
        system_nps_rate_percent=DEFAULT_NPS_RATE * _PERCENT_TO_RATE,
        system_hi_rate_percent=DEFAULT_HI_RATE * _PERCENT_TO_RATE,
        system_ltc_rate_percent=DEFAULT_LTC_RATE_OF_HI * _PERCENT_TO_RATE,
        system_ei_rate_percent=DEFAULT_EI_RATE * _PERCENT_TO_RATE,
    )


async def _get_or_none(db: AsyncSession, client_id: str) -> ClientPayrollDefault | None:
    return (
        await db.execute(
            select(ClientPayrollDefault).where(ClientPayrollDefault.client_id == client_id)
        )
    ).scalar_one_or_none()


async def _authorize_client(db: AsyncSession, client_id: str, user: User) -> Client:
    client = await db.get(Client, client_id)
    if not client or client.tax_office_id != user.tax_office_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Client not found")
    return client


@router.get("/{client_id}/payroll-default", response_model=PayrollDefaultOut)
async def get_payroll_default(
    client_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> PayrollDefaultOut:
    """거래처 세팅값 조회. 없으면 시스템 기본값으로 채워 반환 (lazy creation 안 함)."""
    await _authorize_client(db, client_id, user)
    row = await _get_or_none(db, client_id)
    return _to_response(row) if row else _system_defaults_response()


@router.put("/{client_id}/payroll-default", response_model=PayrollDefaultOut)
async def upsert_payroll_default(
    client_id: str,
    payload: PayrollDefaultUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> PayrollDefaultOut:
    """세팅값 upsert. 백분율 입력은 /100 변환 후 저장."""
    await _authorize_client(db, client_id, user)
    row = await _get_or_none(db, client_id)
    if row is None:
        row = ClientPayrollDefault(client_id=client_id)
        db.add(row)

    patch = payload.model_dump(exclude_unset=True)
    rate_fields = {
        "nps_rate_percent": "nps_rate_override",
        "hi_rate_percent": "hi_rate_override",
        "ltc_rate_percent": "ltc_rate_override",
        "ei_rate_percent": "ei_rate_override",
    }
    for in_field, db_field in rate_fields.items():
        if in_field in patch:
            val = patch.pop(in_field)
            setattr(row, db_field, val / _PERCENT_TO_RATE if val is not None else None)
    for k, v in patch.items():
        setattr(row, k, v)

    await db.commit()
    await db.refresh(row)
    return _to_response(row)


@router.post("/{client_id}/payroll-default/reset", response_model=PayrollDefaultOut)
async def reset_payroll_default(
    client_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> PayrollDefaultOut:
    """시스템 기본값(비과세 한도 + 현행 요율)으로 리셋. 기존 행은 삭제."""
    await _authorize_client(db, client_id, user)
    row = await _get_or_none(db, client_id)
    if row is not None:
        await db.delete(row)
        await db.commit()
    return _system_defaults_response()


@router.post("/{client_id}/employees", response_model=EmployeeOut, status_code=status.HTTP_201_CREATED)
async def create_employee(
    client_id: str,
    payload: EmployeeCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Employee:
    client = await db.get(Client, client_id)
    if not client or client.tax_office_id != user.tax_office_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Client not found")
    rrn_encrypted = encrypt_rrn(payload.rrn) if payload.rrn else None
    rrn_last4 = mask_rrn(payload.rrn).split("-")[-1][:4] if payload.rrn else None
    emp = Employee(
        client_id=client_id,
        name=payload.name,
        rrn_encrypted=rrn_encrypted,
        rrn_last4=rrn_last4,
        employee_code=payload.employee_code,
        hired_at=payload.hired_at,
        status=EmploymentStatus.ACTIVE if payload.rrn else EmploymentStatus.PENDING,
    )
    db.add(emp)
    await db.commit()
    await db.refresh(emp)
    return emp
