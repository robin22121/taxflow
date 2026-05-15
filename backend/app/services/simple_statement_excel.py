"""간이지급명세서 엑셀 생성 — 근로소득 + 사업소득.

국세청 서식 기반:
- 근로소득: 별지 제24호의4서식(1) — 매월 제출
- 사업소득: 별지 제24호의4서식(2) — 매월 제출 (개정 2026.3.20)

위하고T 일괄등록 엑셀과는 별개의 산출물.
세무사가 홈택스에 직접 제출하거나, 위하고T 간이지급명세서 메뉴에서 활용.
"""

from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import logging

from app.models.payroll import IncomeType, PayrollEntry
from app.services.crypto import decrypt_rrn

logger = logging.getLogger(__name__)


def _safe_decrypt_rrn(encrypted: bytes | None) -> str:
    if not encrypted:
        return ""
    try:
        return decrypt_rrn(encrypted)
    except Exception:  # noqa: BLE001
        logger.warning("RRN 복호화 실패 — 마스킹 처리")
        return "******-*******"


# ---------------------------------------------------------------------------
# 근로소득 간이지급명세서
# ---------------------------------------------------------------------------

_WAGE_COLUMNS = [
    "일련번호",
    "귀속연도",
    "귀속월",
    "성명",
    "주민등록번호",
    "근무기간(시작)",
    "근무기간(종료)",
    "급여",
    "상여",
    "비과세소득",
    "소득세",
    "지방소득세",
]


def generate_wage_statement(
    entries: list[PayrollEntry],
    period: str,
) -> bytes:
    """근로소득 간이지급명세서 엑셀 생성.

    Args:
        entries: income_type=WAGE인 PayrollEntry 리스트. employee가 eager-load되어야 함.
        period: "YYYY-MM"
    """
    year, month = period[:4], period[5:7]

    wb = Workbook()
    ws = wb.active
    ws.title = "간이지급명세서(근로소득)"

    _apply_header(ws, _WAGE_COLUMNS)

    for idx, entry in enumerate(entries, start=1):
        emp = entry.employee
        if not emp:
            continue
        rrn = _safe_decrypt_rrn(emp.rrn_encrypted)
        salary = entry.salary_amount if entry.salary_amount is not None else entry.total_amount
        bonus = entry.bonus_amount or 0
        hired = emp.hired_at.isoformat() if emp.hired_at else ""
        resigned = emp.resigned_at.isoformat() if emp.resigned_at else ""

        ws.append([
            idx,
            year,
            month,
            emp.name,
            rrn,
            hired,
            resigned,
            salary,
            bonus,
            entry.non_taxable,
            entry.income_tax,
            entry.local_tax,
        ])

    _apply_widths(ws, [8, 10, 8, 12, 18, 14, 14, 14, 14, 12, 12, 12])
    _add_summary_row(ws, len(entries), len(_WAGE_COLUMNS), sum_cols=[8, 9, 10, 11, 12])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 사업소득 간이지급명세서
# ---------------------------------------------------------------------------

_BUSINESS_COLUMNS = [
    "일련번호",
    "귀속연도",
    "귀속월",
    "업종코드",
    "성명(상호)",
    "주민(사업자)등록번호",
    "외국인여부",
    "지급액",
    "세율(%)",
    "소득세",
    "지방소득세",
    "계",
]


def generate_business_statement(
    entries: list[PayrollEntry],
    period: str,
) -> bytes:
    """사업소득 간이지급명세서 엑셀 생성.

    Args:
        entries: income_type=BUSINESS인 PayrollEntry 리스트. employee가 eager-load되어야 함.
        period: "YYYY-MM"
    """
    year, month = period[:4], period[5:7]

    wb = Workbook()
    ws = wb.active
    ws.title = "간이지급명세서(사업소득)"

    _apply_header(ws, _BUSINESS_COLUMNS)

    for idx, entry in enumerate(entries, start=1):
        emp = entry.employee
        if not emp:
            continue
        rrn = _safe_decrypt_rrn(emp.rrn_encrypted)
        biz_code = entry.business_type_code or emp.business_type_code or "940909"
        # 세율: 기본 3%, 봉사료수취자 5%
        if biz_code == "940905":
            rate = 5
        else:
            rate = 3

        ws.append([
            idx,
            year,
            month,
            biz_code,
            emp.name,
            rrn,
            "",  # 외국인여부
            entry.total_amount,
            rate,
            entry.income_tax,
            entry.local_tax,
            entry.income_tax + entry.local_tax,
        ])

    _apply_widths(ws, [8, 10, 8, 10, 14, 18, 10, 14, 8, 12, 12, 14])
    _add_summary_row(ws, len(entries), len(_BUSINESS_COLUMNS), sum_cols=[8, 10, 11, 12])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _apply_header(ws, columns: list[str]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B4F72")
    center = Alignment(horizontal="center", vertical="center")
    ws.append(columns)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center


def _apply_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_summary_row(ws, entry_count: int, col_count: int, sum_cols: list[int]) -> None:
    """합계 행 추가. sum_cols는 1-indexed 컬럼 번호."""
    summary = [""] * col_count
    summary[0] = "합계"
    for col in sum_cols:
        total = sum(
            ws.cell(row=r, column=col).value or 0
            for r in range(2, entry_count + 2)
        )
        summary[col - 1] = total
    ws.append(summary)
    for col_idx in range(1, col_count + 1):
        ws.cell(row=entry_count + 2, column=col_idx).font = Font(bold=True)
    # 금액 컬럼(=sum_cols) 천단위 콤마 표기 (데이터 행 + 합계 행)
    for col in sum_cols:
        for r in range(2, entry_count + 3):
            ws.cell(row=r, column=col).number_format = "#,##0"


__all__ = [
    "generate_business_statement",
    "generate_wage_statement",
]
