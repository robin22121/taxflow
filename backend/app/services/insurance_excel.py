"""4대 보험 신고서 엑셀 생성 — 자격취득·자격상실·보수월액 변경.

기획안 "4대 보험 업무 자동화" M3(4대 보험 신고서 자동 생성) 실현.
4insure 통합포털 업로드/공단 EDI 활용을 위한 표준 엑셀.
원천세 간이지급명세서(simple_statement_excel)와 별개 산출물.

판정 기준 (기존 모델 필드만 사용 — 스키마 변경 없음):
- 자격취득: match_status=NEW_HIRE_SUSPECTED 또는 입사일이 귀속월에 속함
- 자격상실: match_status=RESIGNATION_SUSPECTED 또는 퇴사일이 귀속월에 속함
- 보수월액 변경: 전월 금액(prev_amount)과 당월 총지급액이 다름 (취득·상실 제외)

통합(combined): 위 3종을 한 워크북 3시트로 — "4대보험 통합" 다운로드용.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.payroll import MatchStatus, PayrollEntry
from app.services.crypto import decrypt_rrn
from app.services.tax_calc import SocialInsurance, calculate_social_insurance

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 정책 상수 — 국민연금 기준소득월액 (2025.7 ~ 2026.6 적용)
# ---------------------------------------------------------------------------
# 출처: 국민연금공단 고시. 보수월액 변경 신고 시 NPS 기준소득월액 유효 범위.
# 주: tax_calc 의 _NPS_MIN/MAX_BASE 는 보험료 산식 캡(별도 정책 캡과 동일치 관리 책임).
_NPS_BASE_MIN = 400_000
_NPS_BASE_MAX = 6_370_000
_NPS_CHANGE_THRESHOLD = 0.20  # 변동률 20% 이상에서 NPS 변경신청 가능

# Employee 모델 미보유 코드 — 본 단계 DB 마이그레이션 비범위.
# 화면/엑셀에서 기본값으로 노출, EDI 단계에서 실제 코드 확정.
_DEFAULT_ACQUISITION_CODE = "01"  # 18세 이상 당연취득
_DEFAULT_LOSS_CODE = "3"  # 사용관계 종료(통상 퇴사)


class _SheetPayload(NamedTuple):
    title: str
    columns: list[str]
    widths: list[int]
    rows: list[list]
    sum_cols: list[int]


def _safe_decrypt_rrn(encrypted: bytes | None) -> str:
    if not encrypted:
        return ""
    try:
        return decrypt_rrn(encrypted)
    except Exception:  # noqa: BLE001
        logger.warning("RRN 복호화 실패 — 마스킹 처리")
        return "******-*******"


def _in_period(d: date | None, period: str) -> bool:
    return d is not None and d.strftime("%Y-%m") == period


# ---------------------------------------------------------------------------
# 분류·판정 — 엑셀 빌더와 JSON 요약이 공유하는 단일 소스
# ---------------------------------------------------------------------------


def _is_acquisition_target(entry: PayrollEntry, period: str) -> bool:
    emp = entry.employee
    if not emp:
        return False
    return (
        entry.match_status == MatchStatus.NEW_HIRE_SUSPECTED
        or _in_period(emp.hired_at, period)
    )


def _is_loss_target(entry: PayrollEntry, period: str) -> bool:
    emp = entry.employee
    if not emp:
        return False
    return (
        entry.match_status == MatchStatus.RESIGNATION_SUSPECTED
        or _in_period(emp.resigned_at, period)
    )


def _is_change_target(entry: PayrollEntry, period: str) -> bool:
    if not entry.employee:
        return False
    if _is_acquisition_target(entry, period) or _is_loss_target(entry, period):
        return False
    return entry.prev_amount is not None and entry.prev_amount != entry.total_amount


@dataclass(frozen=True, slots=True)
class ChangeJudgment:
    """보수월액변경 판정 결과 — 보험별 신청 가능 여부 및 안내."""

    nps_eligible: bool           # 국민연금: 20% 이상 변동
    change_pct: float            # 변동률 (cur-prev)/prev — 부호 보존
    reason_code: str             # "1"=보수인상 "2"=보수인하 (착오정정 3은 판별 불가)
    nps_within_limit: bool       # cur 가 NPS 기준소득월액 한도 내
    nps_consent_required: bool   # NPS 변경신청은 근로자 동의서 필수 (항상 True)


def _change_judgment(entry: PayrollEntry) -> ChangeJudgment:
    prev = entry.prev_amount or 0
    cur = entry.total_amount
    pct = ((cur - prev) / prev) if prev > 0 else 0.0
    nps_eligible = prev > 0 and abs(pct) >= _NPS_CHANGE_THRESHOLD
    within = _NPS_BASE_MIN <= cur <= _NPS_BASE_MAX
    reason = "1" if cur > prev else "2"
    return ChangeJudgment(
        nps_eligible=nps_eligible,
        change_pct=pct,
        reason_code=reason,
        nps_within_limit=within,
        nps_consent_required=True,
    )


# ---------------------------------------------------------------------------
# 화면·외부 공급용 구조 (JSON 직렬화 대상)
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class InsuranceTarget:
    """4대보험 신고 대상자 1명. 자격취득/상실/변경 공용 구조 — 미사용 필드는 None."""

    kind: str                    # "acquisition" | "loss" | "change"
    client_id: str               # 거래처 — 사무소 횡단 화면에서 그룹핑용
    employee_id: str
    name: str
    rrn_last4: str | None        # 마스킹 표시용 (JSON 응답은 절대 전체 RRN 노출 안 함)
    monthly_wage: int            # 보수월액 — 비과세 제외 (= entry.taxable, 없으면 total_amount)
    total_amount: int            # 당월 총지급액
    # 4대보험 사용자부담분 (개별)
    national_pension: int
    health_insurance: int
    longterm_care: int
    employment_insurance: int
    # 자격취득 전용
    hired_at: date | None = None
    acquisition_code: str | None = None
    # 자격상실 전용
    resigned_at: date | None = None
    loss_code: str | None = None
    # 보수월액 변경 전용
    prev_amount: int | None = None
    change_pct: float | None = None
    reason_code: str | None = None
    nps_eligible: bool | None = None
    nps_within_limit: bool | None = None
    nps_consent_required: bool | None = None


@dataclass(frozen=True, slots=True)
class InsuranceSummary:
    period: str
    acquisitions: list[InsuranceTarget]
    losses: list[InsuranceTarget]
    changes: list[InsuranceTarget]


def _monthly_wage(entry: PayrollEntry) -> int:
    """보수월액 — 비과세 제외(=taxable). taxable 미산정 시 total_amount."""
    return entry.taxable if entry.taxable else entry.total_amount


def _si_for(entry: PayrollEntry) -> SocialInsurance:
    """엔트리에 저장된 4대보험 값이 있으면 그대로, 없으면 tax_calc 재계산.

    승인 후 저장된 entry 값이 진실이지만, 신규 자격취득 entry 는 대개 0이므로
    재계산 결과를 fall-back 으로 사용한다.
    """
    if any((entry.national_pension, entry.health_insurance,
            entry.employment_insurance, entry.longterm_care)):
        return SocialInsurance(
            national_pension=entry.national_pension,
            health_insurance=entry.health_insurance,
            employment_insurance=entry.employment_insurance,
            longterm_care=entry.longterm_care,
        )
    return calculate_social_insurance(entry.total_amount, entry.income_type)


def _write_sheet(ws: Worksheet, p: _SheetPayload) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B4F72")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(p.columns)
    for col_idx in range(1, len(p.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in p.rows:
        ws.append(row)

    for i, w in enumerate(p.widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 금액 컬럼(=sum_cols) 천단위 콤마 표기
    n_rows = len(p.rows)
    for col in p.sum_cols:
        for r in range(2, n_rows + 2):
            ws.cell(row=r, column=col).number_format = "#,##0"

    if p.sum_cols and p.rows:
        n = len(p.rows)
        summary: list = [""] * len(p.columns)
        summary[0] = "합계"
        for col in p.sum_cols:
            summary[col - 1] = sum(
                ws.cell(row=r, column=col).value or 0 for r in range(2, n + 2)
            )
        ws.append(summary)
        for col_idx in range(1, len(p.columns) + 1):
            ws.cell(row=n + 2, column=col_idx).font = Font(bold=True)
        for col in p.sum_cols:
            ws.cell(row=n + 2, column=col).number_format = "#,##0"


def _single(p: _SheetPayload) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = p.title
    _write_sheet(ws, p)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 자격취득 신고서
# ---------------------------------------------------------------------------

_ACQUISITION_COLUMNS = [
    "일련번호",
    "성명",
    "주민등록번호",
    "자격취득일",
    "월보수액",
    "국민연금",
    "건강보험",
    "장기요양",
    "고용보험",
    "비고",
]


def _acquisition_payload(entries: list[PayrollEntry], period: str) -> _SheetPayload:
    rows: list[list] = []
    idx = 1
    for entry in entries:
        if not _is_acquisition_target(entry, period):
            continue
        emp = entry.employee  # _is_acquisition_target 이 None 가드 처리
        assert emp is not None
        si = calculate_social_insurance(entry.total_amount, entry.income_type)
        rows.append([
            idx,
            emp.name,
            _safe_decrypt_rrn(emp.rrn_encrypted),
            emp.hired_at.isoformat() if emp.hired_at else "",
            entry.total_amount,
            si.national_pension,
            si.health_insurance,
            si.longterm_care,
            si.employment_insurance,
            "신규취득",
        ])
        idx += 1
    return _SheetPayload(
        "자격취득신고서",
        _ACQUISITION_COLUMNS,
        [8, 12, 18, 14, 14, 12, 12, 12, 12, 12],
        rows,
        [5, 6, 7, 8, 9],
    )


def generate_acquisition_report(entries: list[PayrollEntry], period: str) -> bytes:
    """4대 보험 자격취득 신고서 엑셀 (단일 시트)."""
    return _single(_acquisition_payload(entries, period))


# ---------------------------------------------------------------------------
# 자격상실 신고서
# ---------------------------------------------------------------------------

_LOSS_COLUMNS = [
    "일련번호",
    "성명",
    "주민등록번호",
    "자격상실일",
    "상실부호",
    "당월보수총액",
    "비고",
]


def _loss_payload(entries: list[PayrollEntry], period: str) -> _SheetPayload:
    rows: list[list] = []
    idx = 1
    for entry in entries:
        if not _is_loss_target(entry, period):
            continue
        emp = entry.employee
        assert emp is not None
        rows.append([
            idx,
            emp.name,
            _safe_decrypt_rrn(emp.rrn_encrypted),
            emp.resigned_at.isoformat() if emp.resigned_at else "",
            "3 (퇴직)",
            entry.total_amount,
            "",
        ])
        idx += 1
    return _SheetPayload(
        "자격상실신고서",
        _LOSS_COLUMNS,
        [8, 12, 18, 14, 14, 14, 12],
        rows,
        [6],
    )


def generate_loss_report(entries: list[PayrollEntry], period: str) -> bytes:
    """4대 보험 자격상실 신고서 엑셀 (단일 시트).

    상실부호 3 = 사용관계 종료(퇴직). 실제 사유 분기는 Phase 2 RPA 단계 과제.
    """
    return _single(_loss_payload(entries, period))


# ---------------------------------------------------------------------------
# 보수월액 변경 신고서
# ---------------------------------------------------------------------------

# 웹EDI 보수(소득)월액 변경신청서 — 공식 파일올리기(일괄등록) 레이아웃
# (웹EDI 가입·업무처리 매뉴얼 25.8, 한 건 = 한 행). EDI 업로드용 머신 포맷
# 이므로 천단위 콤마·합계행 미적용(원시 숫자) — research.md §8 참고.
_CHANGE_COLUMNS = [
    "국민연금",            # 신청여부 Y/N
    "건강보험",            # 신청여부 Y/N
    "고용보험",            # 신청여부 Y/N
    "산재보험",            # 신청여부 Y/N
    "성명",
    "주민등록번호",
    "건강증번호",
    "연금-현재소득월액",
    "연금-변경후소득월액",
    "연금-근로자동의",      # 1:동의 2:미동의
    "건강-변경연월",        # YYYYMM
    "건강-보수월액",
    "고용-월평균보수",
    "고용-변경사유",        # 1:보수인상 2:보수인하 3:착오정정
    "산재-월평균보수",
    "산재-변경사유",        # 1:보수인상 2:보수인하 3:착오정정
]


def _change_payload(entries: list[PayrollEntry], period: str) -> _SheetPayload:
    yyyymm = period.replace("-", "")[:6]
    rows: list[list] = []
    for entry in entries:
        if not _is_change_target(entry, period):
            continue
        emp = entry.employee
        assert emp is not None
        prev = entry.prev_amount  # _is_change_target 가 None != cur 보장
        assert prev is not None
        cur = entry.total_amount
        j = _change_judgment(entry)
        nps_flag = "Y" if j.nps_eligible else "N"
        rows.append([
            nps_flag, "Y", "Y", "Y",
            emp.name,
            _safe_decrypt_rrn(emp.rrn_encrypted),
            "",                 # 건강증번호 (미보유)
            prev,               # 연금-현재소득월액
            cur,                # 연금-변경후소득월액
            "1",                # 연금-근로자동의 (세무사 검토 전제 기본 동의)
            yyyymm,             # 건강-변경연월
            cur,                # 건강-보수월액
            cur,                # 고용-월평균보수
            j.reason_code,      # 고용-변경사유
            cur,                # 산재-월평균보수
            j.reason_code,      # 산재-변경사유
        ])
    # EDI 업로드 머신 포맷 → 합계행·콤마 미적용 (sum_cols 비움)
    return _SheetPayload(
        "보수월액변경신고서",
        _CHANGE_COLUMNS,
        [8, 8, 8, 8, 12, 16, 12, 14, 14, 12, 10, 14, 14, 10, 14, 10],
        rows,
        [],
    )


def generate_remuneration_change_report(
    entries: list[PayrollEntry], period: str
) -> bytes:
    """4대 보험 보수월액 변경 신고서 엑셀 (단일 시트)."""
    return _single(_change_payload(entries, period))


# ---------------------------------------------------------------------------
# 통합 (3시트 단일 워크북)
# ---------------------------------------------------------------------------


def generate_combined_insurance_report(
    entries: list[PayrollEntry], period: str
) -> bytes:
    """4대 보험 통합 — 자격취득·자격상실·보수월액변경을 한 워크북 3시트로."""
    payloads = [
        _acquisition_payload(entries, period),
        _loss_payload(entries, period),
        _change_payload(entries, period),
    ]
    wb = Workbook()
    for i, p in enumerate(payloads):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = p.title
        _write_sheet(ws, p)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# JSON 요약 — 화면용 (엑셀과 분류·판정 단일 소스 공유)
# ---------------------------------------------------------------------------


def _acquisition_target(entry: PayrollEntry) -> InsuranceTarget:
    emp = entry.employee
    assert emp is not None
    si = _si_for(entry)
    return InsuranceTarget(
        kind="acquisition",
        client_id=entry.client_id,
        employee_id=emp.id,
        name=emp.name,
        rrn_last4=emp.rrn_last4,
        monthly_wage=_monthly_wage(entry),
        total_amount=entry.total_amount,
        national_pension=si.national_pension,
        health_insurance=si.health_insurance,
        longterm_care=si.longterm_care,
        employment_insurance=si.employment_insurance,
        hired_at=emp.hired_at,
        acquisition_code=_DEFAULT_ACQUISITION_CODE,
    )


def _loss_target(entry: PayrollEntry) -> InsuranceTarget:
    emp = entry.employee
    assert emp is not None
    si = _si_for(entry)
    return InsuranceTarget(
        kind="loss",
        client_id=entry.client_id,
        employee_id=emp.id,
        name=emp.name,
        rrn_last4=emp.rrn_last4,
        monthly_wage=_monthly_wage(entry),
        total_amount=entry.total_amount,
        national_pension=si.national_pension,
        health_insurance=si.health_insurance,
        longterm_care=si.longterm_care,
        employment_insurance=si.employment_insurance,
        resigned_at=emp.resigned_at,
        loss_code=_DEFAULT_LOSS_CODE,
    )


def _change_target(entry: PayrollEntry) -> InsuranceTarget:
    emp = entry.employee
    assert emp is not None
    si = _si_for(entry)
    j = _change_judgment(entry)
    return InsuranceTarget(
        kind="change",
        client_id=entry.client_id,
        employee_id=emp.id,
        name=emp.name,
        rrn_last4=emp.rrn_last4,
        monthly_wage=_monthly_wage(entry),
        total_amount=entry.total_amount,
        national_pension=si.national_pension,
        health_insurance=si.health_insurance,
        longterm_care=si.longterm_care,
        employment_insurance=si.employment_insurance,
        prev_amount=entry.prev_amount,
        change_pct=j.change_pct,
        reason_code=j.reason_code,
        nps_eligible=j.nps_eligible,
        nps_within_limit=j.nps_within_limit,
        nps_consent_required=j.nps_consent_required,
    )


def build_insurance_summary(
    entries: list[PayrollEntry], period: str
) -> InsuranceSummary:
    """엔트리에서 4대보험 신고 대상 3종(자격취득/상실/보수월액변경)을 분류·상세화.

    엑셀 빌더와 동일한 분류 술어(_is_*_target)·판정(_change_judgment)을 사용 — 화면 표기와
    엑셀 출력이 어긋날 수 없음. RRN 은 마지막 4자리만 노출(rrn_last4) — 전체 RRN 미노출.
    """
    acquisitions = [
        _acquisition_target(e) for e in entries if _is_acquisition_target(e, period)
    ]
    losses = [_loss_target(e) for e in entries if _is_loss_target(e, period)]
    changes = [_change_target(e) for e in entries if _is_change_target(e, period)]
    return InsuranceSummary(
        period=period,
        acquisitions=acquisitions,
        losses=losses,
        changes=changes,
    )


__all__ = [
    "ChangeJudgment",
    "InsuranceSummary",
    "InsuranceTarget",
    "build_insurance_summary",
    "generate_acquisition_report",
    "generate_combined_insurance_report",
    "generate_loss_report",
    "generate_remuneration_change_report",
]
