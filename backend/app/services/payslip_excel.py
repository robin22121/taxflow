"""급여(임금)명세서 엑셀 생성 — 근로기준법 제48조.

직원 1명당 시트 1개. 임금 구성항목(기본급·상여·비과세수당) / 공제내역
(소득세·지방소득세·4대보험) / 실수령액을 세로 명세서 형식으로 출력.

기존 PayrollEntry 컬럼만 사용 — 스키마 변경 없음.
RRN은 명세서에 싣지 않음(성명 + 사번으로 식별, PII 노출 최소화).
발송(알림톡/SMS/이메일)은 본 단계 범위 밖 — 다운로드 전용.
"""

from __future__ import annotations

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.models.payroll import PayrollEntry

_INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")

_TITLE_FILL = PatternFill("solid", fgColor="1B4F72")
_SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
_BOLD_WHITE = Font(bold=True, color="FFFFFF", size=13)
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right")


def _safe_sheet_title(name: str, used: set[str]) -> str:
    base = _INVALID_SHEET.sub("", name).strip()[:28] or "직원"
    title = base
    i = 2
    while title in used:
        title = f"{base[:25]}_{i}"
        i += 1
    used.add(title)
    return title


def _basic_pay(entry: PayrollEntry) -> int:
    if entry.salary_amount is not None:
        return entry.salary_amount
    return max(entry.total_amount - (entry.bonus_amount or 0) - entry.non_taxable, 0)


def _write_payslip(ws: Worksheet, entry: PayrollEntry, period: str) -> None:
    emp = entry.employee
    client = entry.client
    client_name = client.business_name if client else ""
    pay_date = entry.payment_date.isoformat() if entry.payment_date else ""

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18

    ws.merge_cells("A1:D1")
    ws["A1"] = "임 금 명 세 서"
    ws["A1"].font = _BOLD_WHITE
    ws["A1"].fill = _TITLE_FILL
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 26

    ws.append(["사업장", client_name, "귀속연월", period])
    ws.append(["성명", emp.name if emp else entry.raw_name,
               "사번", (emp.employee_code if emp else "") or ""])
    ws.append(["임금지급일", pay_date, "", ""])
    for r in (2, 3, 4):
        ws.cell(row=r, column=1).font = _BOLD
        ws.cell(row=r, column=3).font = _BOLD

    meal = entry.meal_amount
    car = entry.car_amount
    childcare = entry.childcare_amount
    other_nontax = max(entry.non_taxable - meal - car - childcare, 0)

    pay_items = [
        ("기본급", _basic_pay(entry)),
        ("상여", entry.bonus_amount or 0),
        ("식대(비과세)", meal),
        ("자가운전보조금(비과세)", car),
        ("육아수당(비과세)", childcare),
        ("기타 비과세", other_nontax),
    ]
    deductions = [
        ("소득세", entry.income_tax),
        ("지방소득세", entry.local_tax),
        ("국민연금", entry.national_pension),
        ("건강보험", entry.health_insurance),
        ("장기요양보험", entry.longterm_care),
        ("고용보험", entry.employment_insurance),
    ]
    deduction_total = sum(v for _, v in deductions)
    net = entry.total_amount - deduction_total

    def _section(header: str) -> None:
        row = ws.max_row + 1
        ws.append([header, "", "", ""])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1)
        c.font = _BOLD
        c.fill = _SECTION_FILL
        c.alignment = _CENTER

    ws.append([])
    _section("지급 항목")
    for label, amount in pay_items:
        ws.append([label, amount, "", ""])
        ws.cell(row=ws.max_row, column=2).number_format = "#,##0"
        ws.cell(row=ws.max_row, column=2).alignment = _RIGHT
    ws.append(["지급액 계", entry.total_amount, "", ""])
    ws.cell(row=ws.max_row, column=1).font = _BOLD
    ws.cell(row=ws.max_row, column=2).font = _BOLD
    ws.cell(row=ws.max_row, column=2).number_format = "#,##0"
    ws.cell(row=ws.max_row, column=2).alignment = _RIGHT

    ws.append([])
    _section("공제 항목")
    for label, amount in deductions:
        ws.append([label, amount, "", ""])
        ws.cell(row=ws.max_row, column=2).number_format = "#,##0"
        ws.cell(row=ws.max_row, column=2).alignment = _RIGHT
    ws.append(["공제액 계", deduction_total, "", ""])
    ws.cell(row=ws.max_row, column=1).font = _BOLD
    ws.cell(row=ws.max_row, column=2).font = _BOLD
    ws.cell(row=ws.max_row, column=2).number_format = "#,##0"
    ws.cell(row=ws.max_row, column=2).alignment = _RIGHT

    ws.append([])
    row = ws.max_row + 1
    ws.append(["실수령액", net, "", ""])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.cell(row=row, column=1).font = _BOLD_WHITE
    ws.cell(row=row, column=1).fill = _TITLE_FILL
    nc = ws.cell(row=row, column=2)
    nc.font = Font(bold=True, size=13)
    nc.number_format = "#,##0"
    nc.alignment = _RIGHT
    ws.cell(row=row, column=1).alignment = _CENTER


def generate_payslips(entries: list[PayrollEntry], period: str) -> bytes:
    """근로기준법 §48 임금명세서 — 직원별 시트 단일 워크북.

    Args:
        entries: WAGE PayrollEntry 리스트. employee, client eager-load 필요.
        period: "YYYY-MM"
    """
    wb = Workbook()
    used: set[str] = set()
    valid = [e for e in entries if e.employee is not None]

    if not valid:
        ws = wb.active
        ws.title = "급여명세서"
        ws["A1"] = "임 금 명 세 서"
        ws["A1"].font = _BOLD_WHITE
        ws["A1"].fill = _TITLE_FILL
        ws["A2"] = f"{period} 대상 근로소득 항목이 없습니다."
        ws.column_dimensions["A"].width = 40
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    for i, entry in enumerate(valid):
        name = entry.employee.name if entry.employee else entry.raw_name
        title = _safe_sheet_title(name, used)
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = title
        _write_payslip(ws, entry, period)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["generate_payslips"]
