"""위하고T 원천세 일괄등록 엑셀 생성.

⚠️ 실제 위하고T 양식의 컬럼 순서·헤더 텍스트·소득구분코드 체계는 양식 실측 이후에 100% 일치시켜야 합니다.
   현재 구현은 일반적인 NTS 원천징수 신고서 양식을 기반으로 한 best-guess이며,
   조명신 사무소에서 실제 양식 1부를 받아 ``WEHAGO_COLUMNS`` 와 ``income_type_to_wehago_code`` 를 정합시켜야 합니다.

사용 예:
    >>> excel_bytes = generate_wehago_excel(entries, period="2026-04")
"""

from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.payroll import IncomeType, PayrollEntry
from app.services.crypto import decrypt_rrn
from app.services.tax_calc import (
    calculate_withholding_tax,
    income_type_to_wehago_code,
)

WEHAGO_COLUMNS: list[str] = [
    "사번",
    "성명",
    "주민등록번호",
    "소득구분코드",
    "지급년월",
    "총지급액",
    "비과세소득",
    "과세소득",
    "소득세",
    "지방소득세",
    "원천징수일자",
]


class WehagoExcelError(ValueError):
    pass


def _ensure_taxes(entry: PayrollEntry) -> tuple[int, int]:
    """If income_tax/local_tax not pre-computed, compute on the fly."""
    if entry.income_tax or entry.local_tax:
        return entry.income_tax, entry.local_tax
    tax = calculate_withholding_tax(
        income_type=entry.income_type,
        taxable_amount=entry.taxable or (entry.total_amount - entry.non_taxable),
        dependents=entry.dependents or 1,
    )
    return tax.income_tax, tax.local_tax


def _row_for_entry(entry: PayrollEntry, period: str) -> list:
    if entry.total_amount <= 0:
        raise WehagoExcelError(f"엔트리 {entry.id}: 총지급액이 0 이하입니다")
    if not entry.employee:
        raise WehagoExcelError(f"엔트리 {entry.id}: 직원 미매칭 상태로 엑셀 생성 불가")

    employee = entry.employee
    rrn = decrypt_rrn(employee.rrn_encrypted) if employee.rrn_encrypted else ""

    income_tax, local_tax = _ensure_taxes(entry)
    taxable = entry.taxable or (entry.total_amount - entry.non_taxable)

    return [
        employee.employee_code or employee.id[:8],
        employee.name,
        rrn,
        income_type_to_wehago_code(entry.income_type),
        period,
        entry.total_amount,
        entry.non_taxable,
        taxable,
        income_tax,
        local_tax,
        entry.payment_date.isoformat() if entry.payment_date else "",
    ]


def generate_wehago_excel(
    entries: Iterable[PayrollEntry],
    period: str,
    sheet_name: str = "원천세_일괄등록",
) -> bytes:
    """PayrollEntry 리스트를 위하고T 일괄등록 표준 엑셀(bytes)로 변환.

    Args:
        entries: 승인된 PayrollEntry 리스트. 직원 객체가 eager-load 되어 있어야 함.
        period: ``"YYYY-MM"`` 형식 지급년월.
        sheet_name: 워크시트 이름.

    Raises:
        WehagoExcelError: 직원 미매칭 또는 금액 이상.
    """
    if not _is_valid_period(period):
        raise WehagoExcelError(f"period 형식이 올바르지 않음: {period!r} (YYYY-MM 필요)")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E5BBA")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(WEHAGO_COLUMNS)
    for col_idx in range(1, len(WEHAGO_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    rows = 0
    income_total = 0
    local_total = 0
    for entry in entries:
        row = _row_for_entry(entry, period)
        ws.append(row)
        rows += 1
        income_total += row[8]
        local_total += row[9]

    if rows == 0:
        raise WehagoExcelError("엔트리가 없어 엑셀을 생성할 수 없습니다")

    # 합계 행
    summary_row = [""] * len(WEHAGO_COLUMNS)
    summary_row[0] = "합계"
    summary_row[5] = sum(e.total_amount for e in entries) if not isinstance(entries, list) else 0
    summary_row[8] = income_total
    summary_row[9] = local_total
    # If entries was a generator, the sum above is wrong — recompute by reading the sheet.
    summary_row[5] = sum(ws.cell(row=r, column=6).value or 0 for r in range(2, rows + 2))
    summary_row[6] = sum(ws.cell(row=r, column=7).value or 0 for r in range(2, rows + 2))
    summary_row[7] = sum(ws.cell(row=r, column=8).value or 0 for r in range(2, rows + 2))
    ws.append(summary_row)

    summary_cell_row = rows + 2
    for col_idx in range(1, len(WEHAGO_COLUMNS) + 1):
        ws.cell(row=summary_cell_row, column=col_idx).font = Font(bold=True)

    # 컬럼 폭 자동
    widths = [10, 12, 18, 12, 12, 14, 12, 14, 12, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _is_valid_period(period: str) -> bool:
    if not period or len(period) != 7 or period[4] != "-":
        return False
    try:
        y = int(period[:4])
        m = int(period[5:])
    except ValueError:
        return False
    return 2000 <= y <= 2100 and 1 <= m <= 12


__all__ = [
    "IncomeType",
    "WEHAGO_COLUMNS",
    "WehagoExcelError",
    "generate_wehago_excel",
]
