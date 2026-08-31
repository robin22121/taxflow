"""급여대장 엑셀 생성 — 위하고T 양식이 아닌 사무소 자체 급여대장 포맷.

양식 참조: 주식회사법인-202604.xlsx
  Row 1: 사원코드 | 사원명 | 부서 | 직급 | 직종 | 수당(merged F~K) | 공제(merged L~W) | 차인지급액
  Row 2:                                     | 기본급 | 상여 | 식대 | 자가운전보조금 | 육아수당 | 지급액계
                                              | 국민연금 | 건강보험 | 고용보험 | 장기요양보험료
                                              | 소득세 | 지방소득세 | 학자금상환액
                                              | 연말정산소득세 | 연말정산지방소득세
                                              | 중도정산소득세 | 중도정산지방소득세 | 공제액계
  Row 3~: data rows
  Last row: 합계
"""

from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.payroll import PayrollEntry
from app.services.tax_calc import calculate_withholding_tax


class PayrollExcelError(ValueError):
    pass


# ── styles ──────────────────────────────────────────────
_HEADER_FONT = Font(name="맑은 고딕", size=9, bold=True)
_DATA_FONT = Font(name="맑은 고딕", size=9)
_HEADER_FILL = PatternFill("solid", fgColor="FFFF99")
_SUM_FONT = Font(name="맑은 고딕", size=9, bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin", color="B3B3B3"),
    right=Side(style="thin", color="B3B3B3"),
    top=Side(style="thin", color="B3B3B3"),
    bottom=Side(style="thin", color="B3B3B3"),
)
_NUM_FMT = "#,##0"

# Column mapping (1-indexed)
# A=사원코드 B=사원명 C=부서 D=직급 E=직종
# F=기본급 G=상여 H=식대 I=자가운전보조금 J=육아수당 K=지급액계
# L=국민연금 M=건강보험 N=고용보험 O=장기요양보험료
# P=소득세 Q=지방소득세 R=학자금상환액
# S=연말정산소득세 T=연말정산지방소득세
# U=중도정산소득세 V=중도정산지방소득세 W=공제액계
# X=차인지급액

COL_COUNT = 24


def _ensure_taxes(entry: PayrollEntry) -> tuple[int, int]:
    if entry.income_tax is not None and entry.local_tax is not None:
        return entry.income_tax, entry.local_tax
    tax = calculate_withholding_tax(
        income_type=entry.income_type,
        taxable_amount=entry.taxable or (entry.total_amount - entry.non_taxable),
        dependents=entry.dependents or 1,
    )
    return tax.income_tax, tax.local_tax


def _build_headers(ws) -> None:
    """Write 2-row merged header matching the 급여대장 template."""
    # Row 1 values
    row1 = {1: "사원코드", 2: "사원명", 3: "부서", 4: "직급", 5: "직종",
            6: "수당", 12: "공제", 24: "차인지급액"}
    # Row 2 values (sub-headers)
    row2 = {6: "기본급", 7: "상여", 8: "식대", 9: "자가운전보조금", 10: "육아수당", 11: "지급액계",
            12: "국민연금", 13: "건강보험", 14: "고용보험", 15: "장기요양보험료",
            16: "소득세", 17: "지방소득세", 18: "학자금상환액",
            19: "연말정산소득세", 20: "연말정산지방소득세",
            21: "중도정산소득세", 22: "중도정산지방소득세", 23: "공제액계"}

    for col, val in row1.items():
        ws.cell(1, col, val)
    for col, val in row2.items():
        ws.cell(2, col, val)

    # Merged ranges
    # A1:A2, B1:B2, C1:C2, D1:D2, E1:E2 — vertical merge for info columns
    for c in range(1, 6):
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)
    # F1:K1 — "수당" header
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=11)
    # L1:W1 — "공제" header
    ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=23)
    # X1:X2 — "차인지급액" vertical merge
    ws.merge_cells(start_row=1, start_column=24, end_row=2, end_column=24)

    # Apply styles to header cells
    for r in range(1, 3):
        for c in range(1, COL_COUNT + 1):
            cell = ws.cell(r, c)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER


def _data_row(entry: PayrollEntry, idx: int) -> list:
    """Build a data row from PayrollEntry."""
    emp = entry.employee
    income_tax, local_tax = _ensure_taxes(entry)

    bonus = entry.bonus_amount or 0
    meal_allowance = entry.meal_amount or 0
    car_allowance = entry.car_amount or 0
    childcare = entry.childcare_amount or 0

    # 기본급 = 총액 - 상여 - 비과세 항목들 (상여·비과세는 총지급액에 이미 포함)
    base_salary = entry.total_amount - bonus - meal_allowance - car_allowance - childcare
    if base_salary < 0:
        base_salary = 0

    # 총지급액은 언제나 total_amount (상여·비과세는 그 내부 분해이므로 재가산 금지)
    gross = entry.total_amount

    national_pension = entry.national_pension or 0
    health_insurance = entry.health_insurance or 0
    employment_insurance = entry.employment_insurance or 0
    longterm_care = entry.longterm_care or 0

    deduction_total = (national_pension + health_insurance + employment_insurance
                       + longterm_care + income_tax + local_tax)
    net_pay = gross - deduction_total

    emp_code = (emp.employee_code if emp else None) or str(idx)
    emp_name = (emp.name if emp else None) or entry.raw_name

    return [
        emp_code,                             # A: 사원코드
        emp_name,                             # B: 사원명
        "",                                   # C: 부서
        "",                                   # D: 직급
        "",                                   # E: 직종
        base_salary,                          # F: 기본급
        bonus,                                # G: 상여
        meal_allowance,                       # H: 식대
        car_allowance,                        # I: 자가운전보조금
        childcare,                            # J: 육아수당
        gross,                                # K: 지급액계
        national_pension,                     # L: 국민연금
        health_insurance,                     # M: 건강보험
        employment_insurance,                 # N: 고용보험
        longterm_care,                        # O: 장기요양보험료
        income_tax,                           # P: 소득세
        local_tax,                            # Q: 지방소득세
        0,                                    # R: 학자금상환액
        0,                                    # S: 연말정산소득세
        0,                                    # T: 연말정산지방소득세
        0,                                    # U: 중도정산소득세
        0,                                    # V: 중도정산지방소득세
        deduction_total,                      # W: 공제액계
        net_pay,                              # X: 차인지급액
    ]


def generate_payroll_excel(
    entries: Iterable[PayrollEntry],
    period: str,
    client_name: str = "",
) -> bytes:
    """PayrollEntry 리스트를 급여대장 엑셀(bytes)로 변환."""
    if not _is_valid_period(period):
        raise PayrollExcelError(f"period 형식이 올바르지 않음: {period!r} (YYYY-MM 필요)")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{client_name}-{period}" if client_name else f"급여대장-{period}"

    _build_headers(ws)

    data_start = 3
    row_count = 0
    for idx, entry in enumerate(entries, start=1):
        row_data = _data_row(entry, idx)
        ws.append(row_data)
        r = data_start + row_count
        for c in range(1, COL_COUNT + 1):
            cell = ws.cell(r, c)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            if c >= 6:
                cell.number_format = _NUM_FMT
                cell.alignment = _RIGHT
            else:
                cell.alignment = _CENTER
        row_count += 1

    if row_count == 0:
        raise PayrollExcelError("엔트리가 없어 엑셀을 생성할 수 없습니다")

    # 합계 row
    sum_row = data_start + row_count
    ws.cell(sum_row, 1, "합계")
    ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=5)

    for c in range(6, COL_COUNT + 1):
        col_letter = get_column_letter(c)
        ws.cell(sum_row, c).value = f"=SUM({col_letter}{data_start}:{col_letter}{sum_row - 1})"
        ws.cell(sum_row, c).number_format = _NUM_FMT
        ws.cell(sum_row, c).alignment = _RIGHT

    for c in range(1, COL_COUNT + 1):
        ws.cell(sum_row, c).font = _SUM_FONT
        ws.cell(sum_row, c).border = _THIN_BORDER

    # Column widths
    for c in range(1, COL_COUNT + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14 if c == 1 else 13

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _is_valid_period(period: str) -> bool:
    if not period or len(period) != 7 or period[4] != "-":
        return False
    try:
        y, m = int(period[:4]), int(period[5:])
    except ValueError:
        return False
    return 2000 <= y <= 2100 and 1 <= m <= 12


__all__ = ["PayrollExcelError", "generate_payroll_excel"]
