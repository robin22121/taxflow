"""Convert uploaded files into the plain-text representation our AI parser expects.

Supported types in Phase 1:
- Audio (.mp3 / .m4a / .wav / .aac) — put to Object Storage, presign URL, send to STT.
- Excel (.xlsx / .xls) — read cells with openpyxl, build a tab-separated text dump.
- CSV (.csv) — decoded as text directly.
- Image (.png / .jpg / .jpeg / .webp) — pass-through to Claude Vision via the
  ``images`` list on ``IntakeResult``.
- PDF (.pdf) — render each page to PNG via pypdfium2 and pass to Vision (max 20 pages).
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from typing import Iterable

from openpyxl import load_workbook

from app.services.pii import redact_pii
from app.services.storage import ObjectStorage

logger = logging.getLogger(__name__)


AUDIO_EXTS = (".mp3", ".m4a", ".wav", ".aac", ".ogg", ".flac")
EXCEL_EXTS = (".xlsx", ".xlsm", ".xls")
CSV_EXTS = (".csv",)
TEXT_EXTS = (".txt",)
IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".webp")
PDF_EXTS = (".pdf",)

# Vision 비용·응답 시간 보호 — 너무 큰 PDF는 잘라서 처리
_MAX_PDF_PAGES = 20
_PDF_RENDER_DPI = 150


@dataclass(slots=True)
class IntakeResult:
    text: str
    kind: str  # "audio" | "excel" | "csv" | "image" | "pdf" | "unknown"
    storage_key: str | None = None
    note: str | None = None
    # Vision 모델로 보낼 (bytes, mime) 리스트. 이미지는 1개, PDF는 페이지 수만큼.
    images: list[tuple[bytes, str]] = field(default_factory=list)


def _is_audio(filename: str) -> bool:
    return filename.lower().endswith(AUDIO_EXTS)


def _is_excel(filename: str) -> bool:
    return filename.lower().endswith(EXCEL_EXTS)


def _is_csv(filename: str) -> bool:
    return filename.lower().endswith(CSV_EXTS)


def _is_text(filename: str) -> bool:
    return filename.lower().endswith(TEXT_EXTS)


def _is_image(filename: str) -> bool:
    return filename.lower().endswith(IMAGE_EXTS)


def _is_pdf(filename: str) -> bool:
    return filename.lower().endswith(PDF_EXTS)


def _pdf_to_page_images(blob: bytes) -> list[tuple[bytes, str]]:
    """Render each PDF page to PNG bytes. Returns up to _MAX_PDF_PAGES."""
    import pypdfium2 as pdfium

    pdf = pdfium.PdfDocument(blob)
    pages: list[tuple[bytes, str]] = []
    scale = _PDF_RENDER_DPI / 72  # PDF default DPI = 72
    for i in range(min(len(pdf), _MAX_PDF_PAGES)):
        page = pdf[i]
        pil_image = page.render(scale=scale).to_pil()
        buf = io.BytesIO()
        pil_image.save(buf, format="PNG")
        pages.append((buf.getvalue(), "image/png"))
    if len(pdf) > _MAX_PDF_PAGES:
        logger.warning(
            "[file-intake] PDF has %d pages, only processing first %d",
            len(pdf), _MAX_PDF_PAGES,
        )
    return pages


def _excel_to_text(blob: bytes) -> str:
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    out_lines: list[str] = []
    for ws in wb.worksheets:
        out_lines.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(cells):
                out_lines.append("\t".join(cells))
    return "\n".join(out_lines)


def _csv_to_text(blob: bytes) -> str:
    text = blob.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    return "\n".join("\t".join(row) for row in rows if any(c.strip() for c in row))


async def intake_file(
    *,
    filename: str,
    content: bytes,
    storage: ObjectStorage,
) -> IntakeResult:
    """Turn an uploaded file into plain text suitable for the AI parser."""
    if _is_audio(filename):
        # 음성/통화 처리 미지원 (정책 폐기) — 파일 보관 없이 안내만 반환
        return IntakeResult(
            text="",
            kind="audio",
            note="음성 파일은 현재 처리하지 않습니다.",
        )

    if _is_excel(filename):
        ext = "." + filename.rsplit(".", 1)[-1].lower()
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if ext == ".xlsx" else "application/vnd.ms-excel"
        key = storage.make_key("excel", ext)
        storage.put_object(key, content, content_type=mime)
        return IntakeResult(
            text=redact_pii(_excel_to_text(content)),
            kind="excel",
            storage_key=key,
        )

    if _is_csv(filename):
        key = storage.make_key("csv", ".csv")
        storage.put_object(key, content, content_type="text/csv")
        return IntakeResult(
            text=redact_pii(_csv_to_text(content)),
            kind="csv",
            storage_key=key,
        )

    if _is_text(filename):
        key = storage.make_key("text", ".txt")
        storage.put_object(key, content, content_type="text/plain")
        text = content.decode("utf-8-sig", errors="replace")
        return IntakeResult(
            text=redact_pii(text),
            kind="text",
            storage_key=key,
        )

    if _is_image(filename):
        # 이미지는 Object Storage에 보존(감사용) + raw bytes를 함께 반환해
        # AI 파서가 Vision 모델로 직접 분석하도록 한다.
        ext = "." + filename.rsplit(".", 1)[-1].lower()
        mime = _image_mime(ext)
        key = storage.make_key("image", ext)
        storage.put_object(key, content, content_type=mime)
        return IntakeResult(
            text=f"[이미지 첨부: {filename}]",  # AI에 컨텍스트만 제공 — 실제 내용은 images로
            kind="image",
            storage_key=key,
            images=[(content, mime)],
        )

    if _is_pdf(filename):
        # PDF는 원본 보존 + 각 페이지를 PNG로 변환해 Vision 으로 보낸다.
        key = storage.make_key("pdf", ".pdf")
        storage.put_object(key, content, content_type="application/pdf")
        try:
            pages = _pdf_to_page_images(content)
        except Exception as e:  # noqa: BLE001
            logger.exception("[file-intake] PDF parse failed: %s", filename)
            return IntakeResult(
                text=f"[PDF 첨부: {filename} — 페이지 변환 실패]",
                kind="pdf",
                storage_key=key,
                note=f"PDF 변환 오류: {e}",
            )
        if not pages:
            return IntakeResult(
                text=f"[PDF 첨부: {filename} — 페이지 없음]",
                kind="pdf",
                storage_key=key,
                note="빈 PDF",
            )
        return IntakeResult(
            text=f"[PDF 첨부: {filename}, {len(pages)}페이지]",
            kind="pdf",
            storage_key=key,
            images=pages,
        )

    return IntakeResult(text="", kind="unknown", note=f"지원하지 않는 파일 형식: {filename}")


def _image_mime(ext: str) -> str:
    return {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }.get(ext, "application/octet-stream")


__all__ = ["IntakeResult", "intake_file"]
