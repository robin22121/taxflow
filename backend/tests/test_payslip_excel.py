"""급여(임금)명세서 엑셀 생성기 테스트 (in-memory, no DB)."""

from datetime import date
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.payslip_excel import generate_payslips


def _entry(name, *, code="E001", client="(주)에이상사", salary=3_000_000,
           bonus=0, non_tax=200_000, meal=200_000, car=0, childcare=0,
           income_tax=85_000, local_tax=8_500, nps=135_000, hi=106_350,
           ltc=13_770, ei=27_000, total=3_200_000):
    return SimpleNamespace(
        raw_name=name,
        employee=SimpleNamespace(name=name, employee_code=code),
        client=SimpleNamespace(business_name=client),
        payment_date=date(2026, 4, 25),
        salary_amount=salary,
        bonus_amount=bonus,
        non_taxable=non_tax,
        meal_amount=meal,
        car_amount=car,
        childcare_amount=childcare,
        income_tax=income_tax,
        local_tax=local_tax,
        national_pension=nps,
        health_insurance=hi,
        longterm_care=ltc,
        employment_insurance=ei,
        total_amount=total,
    )


def _cells(ws):
    return {(c.row, c.column): c.value for row in ws.iter_rows() for c in row}


def test_sheet_per_employee_and_net_pay():
    e1 = _entry("김연호")
    e2 = _entry("박민수", code="E002")
    wb = load_workbook(BytesIO(generate_payslips([e1, e2], "2026-04")))

    assert wb.sheetnames == ["김연호", "박민수"]
    ws = wb["김연호"]
    flat = list(ws.values)
    # 제목
    assert ws["A1"].value == "임 금 명 세 서"
    # 실수령액 = 지급계 - 공제계 (85000+8500+135000+106350+13770+27000=375620)
    net_row = next(r for r in flat if r and r[0] == "실수령액")
    assert net_row[1] == 3_200_000 - 375_620


def test_duplicate_names_get_unique_sheets():
    wb = load_workbook(BytesIO(generate_payslips(
        [_entry("홍길동"), _entry("홍길동", code="E099")], "2026-04")))
    assert len(wb.sheetnames) == 2
    assert wb.sheetnames[0] == "홍길동"
    assert wb.sheetnames[1] != "홍길동"


def test_empty_returns_valid_workbook():
    wb = load_workbook(BytesIO(generate_payslips([], "2026-04")))
    assert wb.active["A1"].value == "임 금 명 세 서"
