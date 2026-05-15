"""4대 보험 신고서 엑셀 생성기 테스트 (in-memory, no DB)."""

from datetime import date
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.models.employee import EmploymentStatus
from app.models.payroll import IncomeType, MatchStatus
from app.services.crypto import encrypt_rrn
from app.services.insurance_excel import (
    generate_acquisition_report,
    generate_combined_insurance_report,
    generate_loss_report,
    generate_remuneration_change_report,
)


def _emp(name, rrn="900101-1234567", hired=None, resigned=None,
         status=EmploymentStatus.ACTIVE):
    return SimpleNamespace(
        id="e_" + name,
        name=name,
        rrn_encrypted=encrypt_rrn(rrn) if rrn else None,
        hired_at=hired,
        resigned_at=resigned,
        status=status,
    )


def _entry(emp, total, *, match=MatchStatus.MATCHED, prev=None,
           income_type=IncomeType.WAGE):
    return SimpleNamespace(
        id="p_" + emp.name,
        employee=emp,
        income_type=income_type,
        total_amount=total,
        match_status=match,
        prev_amount=prev,
    )


def _rows(blob: bytes):
    ws = load_workbook(BytesIO(blob)).active
    return [[c.value for c in row] for row in ws.iter_rows()]


def test_acquisition_includes_new_hire_and_period_join():
    new_by_status = _entry(_emp("김신규"), 3_000_000,
                            match=MatchStatus.NEW_HIRE_SUSPECTED)
    new_by_hire = _entry(_emp("이입사", hired=date(2026, 4, 12)), 2_500_000)
    existing = _entry(_emp("박기존", hired=date(2024, 1, 1)), 2_800_000)

    rows = _rows(generate_acquisition_report(
        [new_by_status, new_by_hire, existing], "2026-04"))
    names = {r[1] for r in rows[1:-1]}  # skip header + 합계
    assert names == {"김신규", "이입사"}
    assert rows[0][0] == "일련번호"
    # 국민연금 컬럼(6번째)이 0보다 큼 — calculate_social_insurance 연동 확인
    assert rows[1][5] > 0


def test_loss_includes_resignation():
    resigned = _entry(
        _emp("최퇴사", resigned=date(2026, 4, 20),
             status=EmploymentStatus.RESIGNED),
        2_000_000,
    )
    suspected = _entry(_emp("정의심"), 1_900_000,
                       match=MatchStatus.RESIGNATION_SUSPECTED)
    active = _entry(_emp("한재직"), 2_100_000)

    rows = _rows(generate_loss_report([resigned, suspected, active], "2026-04"))
    names = {r[1] for r in rows[1:-1]}
    assert names == {"최퇴사", "정의심"}


def test_remuneration_change_only_changed_amounts():
    changed = _entry(_emp("오변동", hired=date(2023, 5, 1)),
                     3_300_000, prev=3_000_000)
    same = _entry(_emp("강동결", hired=date(2023, 5, 1)),
                  3_000_000, prev=3_000_000)
    new_hire = _entry(_emp("신규자", hired=date(2026, 4, 1)),
                      2_000_000, prev=None)

    rows = _rows(generate_remuneration_change_report(
        [changed, same, new_hire], "2026-04"))
    data = rows[1:-1]
    assert {r[1] for r in data} == {"오변동"}
    # 증감 = 변경후 - 변경전
    assert data[0][5] == 300_000


def test_empty_returns_valid_workbook():
    blob = generate_acquisition_report([], "2026-04")
    ws = load_workbook(BytesIO(blob)).active
    assert [c.value for c in ws[1]][0] == "일련번호"


def test_combined_has_three_sheets():
    acq = _entry(_emp("김신규"), 3_000_000, match=MatchStatus.NEW_HIRE_SUSPECTED)
    loss = _entry(_emp("최퇴사", resigned=date(2026, 4, 20)), 2_000_000,
                  match=MatchStatus.RESIGNATION_SUSPECTED)
    chg = _entry(_emp("오변동", hired=date(2023, 1, 1)), 3_300_000, prev=3_000_000)

    wb = load_workbook(BytesIO(
        generate_combined_insurance_report([acq, loss, chg], "2026-04")))
    assert wb.sheetnames == ["자격취득신고서", "자격상실신고서", "보수월액변경신고서"]
    assert wb["자격취득신고서"]["B2"].value == "김신규"
    assert wb["자격상실신고서"]["B2"].value == "최퇴사"
    assert wb["보수월액변경신고서"]["B2"].value == "오변동"
