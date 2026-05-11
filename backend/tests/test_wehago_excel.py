"""Wehago Excel generator tests (in-memory, no DB)."""

from datetime import date
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.models.payroll import IncomeType
from app.services.crypto import encrypt_rrn
from app.services.wehago_excel import (
    WEHAGO_COLUMNS,
    WehagoExcelError,
    generate_wehago_excel,
)


def _fake_employee(name: str, rrn: str, code: str = ""):
    return SimpleNamespace(
        id="e_" + name,
        name=name,
        employee_code=code or None,
        rrn_encrypted=encrypt_rrn(rrn) if rrn else None,
    )


def _fake_entry(
    name: str,
    total: int,
    income_tax: int = 0,
    local_tax: int = 0,
    income_type: IncomeType = IncomeType.WAGE,
    rrn: str = "900101-1234567",
):
    emp = _fake_employee(name, rrn)
    return SimpleNamespace(
        id="p_" + name,
        employee=emp,
        income_type=income_type,
        total_amount=total,
        non_taxable=0,
        taxable=total,
        income_tax=income_tax,
        local_tax=local_tax,
        payment_date=date(2026, 4, 30),
        dependents=1,
    )


def test_generates_valid_xlsx_bytes():
    entries = [
        _fake_entry("김연호", 3_000_000, income_tax=85_000, local_tax=8_500),
        _fake_entry("박민수", 2_500_000, income_tax=46_000, local_tax=4_600),
    ]
    blob = generate_wehago_excel(entries, period="2026-04")
    wb = load_workbook(BytesIO(blob))
    ws = wb.active

    headers = [c.value for c in ws[1]]
    assert headers == WEHAGO_COLUMNS

    # row 2 = 김연호, row 3 = 박민수, row 4 = 합계
    assert ws.cell(row=2, column=2).value == "김연호"
    assert ws.cell(row=2, column=6).value == 3_000_000
    assert ws.cell(row=2, column=9).value == 85_000

    summary = ws.cell(row=4, column=1).value
    assert summary == "합계"
    assert ws.cell(row=4, column=6).value == 5_500_000
    assert ws.cell(row=4, column=9).value == 131_000


def test_invalid_period_raises():
    with pytest.raises(WehagoExcelError):
        generate_wehago_excel(
            [_fake_entry("김연호", 1_000_000, income_tax=10_000, local_tax=1_000)],
            period="2026-13",
        )


def test_empty_entries_raises():
    with pytest.raises(WehagoExcelError):
        generate_wehago_excel([], period="2026-04")


def test_missing_employee_raises():
    e = _fake_entry("test", 1_000_000)
    e.employee = None
    with pytest.raises(WehagoExcelError):
        generate_wehago_excel([e], period="2026-04")


def test_auto_calculates_when_tax_zero():
    # income_tax=None → 자동 계산
    entry = _fake_entry("김연호", 3_000_000, income_tax=None, local_tax=None)
    blob = generate_wehago_excel([entry], period="2026-04")
    wb = load_workbook(BytesIO(blob))
    ws = wb.active
    # 자동 계산된 income_tax는 0보다 커야 함 (300만원 근로소득)
    assert ws.cell(row=2, column=9).value > 30_000


def test_business_income_3_3_percent_in_excel():
    entry = _fake_entry(
        "이프리",
        1_000_000,
        income_tax=None,  # 자동 계산
        local_tax=None,
        income_type=IncomeType.BUSINESS,
    )
    blob = generate_wehago_excel([entry], period="2026-04")
    wb = load_workbook(BytesIO(blob))
    ws = wb.active
    assert ws.cell(row=2, column=4).value == "A03"  # 사업소득 코드
    assert ws.cell(row=2, column=9).value == 30_000
    assert ws.cell(row=2, column=10).value == 3_000
